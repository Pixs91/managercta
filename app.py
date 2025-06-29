from flask import Flask, render_template, request, redirect, url_for, send_from_directory
from flask_login import LoginManager, login_user, login_required, logout_user, UserMixin
import pandas as pd
import unicodedata
import re
import os
import json
from datetime import datetime
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
EXPECTED_COLUMNS = {
    'bolt': {'șofer', 'câștiguri nete|lei', 'numerar încasat|lei'},
    'uber': {'prenumele șoferului', 'numele de familie al șoferului',
             'câștiguri primite : câștigurile tale',
             'câștiguri primite : sold cursă : plăți : numerar încasat'}
}
app = Flask(__name__)
app.secret_key = 'secret-key-123'

login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

class User(UserMixin):
    def __init__(self, id):
        self.id = id
        self.password = "firstclassauto"

users = {"george": User(id="george")}

@login_manager.user_loader
def load_user(user_id):
    return users.get(user_id)
def validate_columns(df, expected, source):
    actual = set(df.columns.str.strip().str.lower().str.replace('\xa0', ' '))
    missing = expected - actual
    if missing:
        raise ValueError(f"❌ Fișierul {source.upper()} lipsesc coloanele: {', '.join(missing)}")
def normalize_name(name):
    name = unicodedata.normalize('NFKD', str(name)).encode('ascii', 'ignore').decode()
    name = name.lower().strip().replace('-', ' ')
    name = re.sub(r'\s+', ' ', name)
    return ' '.join(sorted(name.split())).title()

def clean_columns(df, source):
    df.columns = df.columns.str.strip().str.lower().str.replace('\xa0', ' ')
    if source == 'bolt':
        df.rename(columns={
            'șofer': 'driver',
            'câștiguri nete|lei': 'bolt_net',
            'numerar încasat|lei': 'bolt_cash'
        }, inplace=True)
        df['driver'] = df['driver'].apply(normalize_name)
        return df[['driver', 'bolt_net', 'bolt_cash']]
    elif source == 'uber':
        df['driver'] = (
            df['prenumele șoferului'].fillna('') + ' ' +
            df['numele de familie al șoferului'].fillna('')
        ).apply(normalize_name)
        df.rename(columns={
            'câștiguri primite : câștigurile tale': 'uber_net',
            'câștiguri primite : sold cursă : plăți : numerar încasat': 'uber_cash'
        }, inplace=True)
        df['uber_cash'] = df['uber_cash'].abs()
        return df[['driver', 'uber_net', 'uber_cash']]

def save_history(df, week, top_driver=None, top_earning=None):
    folder = os.path.join('uploads', 'history', week)
    os.makedirs(folder, exist_ok=True)

    column_order = [
        'driver', 'bolt_net', 'uber_net', 'total_net', 'commission',
        'after_commission', 'service_fee', 'bolt_cash', 'uber_cash',
        'total_cash', 'to_be_paid'
    ]

    column_names = {
        'driver': 'Șofer',
        'bolt_net': 'Câștig Net Bolt',
        'uber_net': 'Câștig Net Uber',
        'total_net': 'Total Net',
        'commission': 'Comision 9%',
        'after_commission': 'După Comision',
        'service_fee': 'Taxă de Serviciu',
        'bolt_cash': 'Numerar Bolt',
        'uber_cash': 'Numerar Uber',
        'total_cash': 'Numerar Total',
        'to_be_paid': 'De Plătit'
    }

    export_df = df[column_order].rename(columns=column_names)

# Clone for Excel output and append TOTAL row
    excel_df = export_df.copy()
    totals = {col: 0.0 for col in export_df.columns}
    totals['Șofer'] = 'TOTAL'
    totals['Câștig Net Bolt'] = export_df['Câștig Net Bolt'].sum()
    totals['Câștig Net Uber'] = export_df['Câștig Net Uber'].sum()
    totals['Numerar Bolt'] = export_df['Numerar Bolt'].sum()
    totals['Numerar Uber'] = export_df['Numerar Uber'].sum()
    excel_df = pd.concat([excel_df, pd.DataFrame([totals])], ignore_index=True)


    path = os.path.join(folder, 'summary.xlsx')
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='Rezumat')
        sheet = writer.sheets['Rezumat']

        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
        align = Alignment(horizontal='center')

        for col, name in enumerate(export_df.columns, 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align
            sheet.column_dimensions[get_column_letter(col)].width = max(14, len(name) + 4)

        total_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        last_row = export_df.shape[0] + 1
        for col in range(1, len(export_df.columns) + 1):
            cell = sheet.cell(row=last_row, column=col)
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = align

        sheet.freeze_panes = 'A2'

    metadata = {
        "week": week,
        "timestamp": datetime.now().isoformat(),
        "summary_file": f"/download_history/{week}/summary.xlsx"
    }
    if top_driver and top_earning:
        metadata["top_driver"] = top_driver
        metadata["top_earning"] = top_earning

    with open(os.path.join(folder, 'metadata.json'), 'w') as f:
        json.dump(metadata, f)

def export_pending_platform_sums(df, week):
    folder = os.path.join('uploads', 'history', week)
    os.makedirs(folder, exist_ok=True)

    bolt_net = df['bolt_net'].sum()
    bolt_cash = df['bolt_cash'].sum()
    uber_net = df['uber_net'].sum()
    uber_cash = df['uber_cash'].sum()

    bolt_due = bolt_net - bolt_cash
    uber_due = uber_net - uber_cash
    total_due = bolt_due + uber_due

    wb = Workbook()
    ws = wb.active
    ws.title = "Sume Platforme"

    ws.merge_cells('C1:G1')
    ws['C1'] = "SUME AȘTEPTATE DIN PARTEA PLATFORMELOR"
    ws['C1'].font = Font(size=16, bold=True)
    ws['C1'].alignment = Alignment(horizontal='center')

# Bolt
    ws[f'E3'] = "Bolt"
    ws[f'E3'].font = Font(size=13, bold=True, color="228B22")
    ws[f'E3'].alignment = Alignment(horizontal='center')

    ws[f'E4'] = f"{bolt_due:.2f} RON"
    ws[f'E4'].font = Font(size=14, bold=True, color="228B22")
    ws[f'E4'].alignment = Alignment(horizontal='center')

# Uber
    ws[f'E6'] = "Uber"
    ws[f'E6'].font = Font(size=13, bold=True, color="000000")
    ws[f'E6'].alignment = Alignment(horizontal='center')

    ws[f'E7'] = f"{uber_due:.2f} RON"
    ws[f'E7'].font = Font(size=14, bold=True, color="000000")
    ws[f'E7'].alignment = Alignment(horizontal='center')

# Total
    ws[f'E9'] = "TOTAL VENIT NET"
    ws[f'E9'].font = Font(size=13, bold=True)
    ws[f'E9'].alignment = Alignment(horizontal='center')

    ws[f'E10'] = f"{total_due:.2f} RON"
    ws[f'E10'].font = Font(size=15, bold=True)
    ws[f'E10'].alignment = Alignment(horizontal='center')


    wb.save(os.path.join(folder, 'sume_platforme.xlsx'))
def generate_driver_reports(df, week):
    folder = os.path.join('uploads', 'history', week, 'drivers')
    os.makedirs(folder, exist_ok=True)

    column_order = [
        'driver', 'bolt_net', 'uber_net', 'total_net', 'commission',
        'after_commission', 'service_fee', 'bolt_cash', 'uber_cash',
        'total_cash', 'to_be_paid'
    ]

    column_names = {
        'driver': 'Șofer',
        'bolt_net': 'Câștig Net Bolt',
        'uber_net': 'Câștig Net Uber',
        'total_net': 'Total Net',
        'commission': 'Comision 9%',
        'after_commission': 'După Comision',
        'service_fee': 'Taxă de Serviciu',
        'bolt_cash': 'Numerar Bolt',
        'uber_cash': 'Numerar Uber',
        'total_cash': 'Numerar Total',
        'to_be_paid': 'De Plătit'
    }

    for _, row in df.iterrows():
        export_df = pd.DataFrame([row])[column_order].rename(columns=column_names)
        path = os.path.join(folder, f"{row['driver']}.xlsx")

        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='Raport')
            sheet = writer.sheets['Raport']

            font = Font(bold=True, size=12, color="FFFFFF")
            fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
            align = Alignment(horizontal='center')

            for col_num, header in enumerate(export_df.columns, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = font
                cell.fill = fill
                cell.alignment = align
                sheet.column_dimensions[get_column_letter(col_num)].width = max(14, len(header) + 4)

    return folder

def zip_driver_reports(folder, week):
    zip_path = os.path.join('uploads', 'history', week, 'drivers.zip')
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file in os.listdir(folder):
            zipf.write(os.path.join(folder, file), arcname=file)
    return zip_path

# ---------------- Flask Routes ----------------

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u = request.form.get('username').lower()
        p = request.form.get('password')
        user = users.get(u)
        if user and user.password == p:
            login_user(user)
            return redirect(url_for('index'))
        return render_template('login.html', error="Invalid credentials")
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/upload', methods=['POST'])
@login_required
def upload():
    week = request.form['week']
    bolt_files = request.files.getlist('bolt')
    bolt_raw = pd.concat([pd.read_excel(f) for f in bolt_files], ignore_index=True)
    print("\n🧾 RAW BOLT COLUMNS (as Python sees them):")
    print(list(bolt_raw.columns))

    validate_columns(bolt_raw, EXPECTED_COLUMNS['bolt'], 'bolt')
    bolt_df = clean_columns(bolt_raw, 'bolt')

    uber_files = request.files.getlist('uber')
    uber_raw = pd.concat([pd.read_excel(f) for f in uber_files], ignore_index=True)
    validate_columns(uber_raw, EXPECTED_COLUMNS['uber'], 'uber')
    uber_df = clean_columns(uber_raw, 'uber')

    combined = pd.concat([bolt_df, uber_df], ignore_index=True)
    df = combined.groupby('driver', as_index=False).sum(numeric_only=True).fillna(0)

    for col in ['bolt_net', 'bolt_cash', 'uber_net', 'uber_cash']:
        if col not in df.columns:
            df[col] = 0

    df['total_net'] = df['bolt_net'] + df['uber_net']
    df['commission'] = df['total_net'] * 0.09
    df['after_commission'] = df['total_net'] - df['commission']
    df['service_fee'] = 30
    df['total_cash'] = df['bolt_cash'] + df['uber_cash']
    df['to_be_paid'] = df['after_commission'] - df['service_fee'] - df['total_cash']
    df['final_gross'] = df['to_be_paid'] + df['total_cash']

    top3 = df.nlargest(3, 'final_gross')[['driver', 'to_be_paid', 'total_cash', 'final_gross']].copy()
    top3['final_gross_display'] = top3['final_gross'].map(lambda x: f"{x:,.2f} RON")

    top_driver = top3.iloc[0]['driver']
    top_earning = top3.iloc[0]['final_gross_display']

    save_history(df, week, top_driver, top_earning)
    folder = generate_driver_reports(df, week)
    zip_driver_reports(folder, week)
    export_pending_platform_sums(df, week)

    return render_template('summary.html',
                           week=week,
                           summary=df.to_dict(orient='records'),
                           top3=top3.to_dict(orient='records'))
@app.route('/summary/<week>')
@login_required
def view_summary(week):
    file = os.path.join('uploads', 'history', week, 'summary.xlsx')
    if not os.path.exists(file):
        return f"Summary file for {week} not found", 404

    df = pd.read_excel(file)

    rename_map = {
        'Șofer': 'driver',
        'Câștig Net Bolt': 'bolt_net',
        'Câștig Net Uber': 'uber_net',
        'Total Net': 'total_net',
        'Comision 9%': 'commission',
        'După Comision': 'after_commission',
        'Taxă de Serviciu': 'service_fee',
        'Numerar Bolt': 'bolt_cash',
        'Numerar Uber': 'uber_cash',
        'Numerar Total': 'total_cash',
        'De Plătit': 'to_be_paid'
    }
    df.rename(columns=rename_map, inplace=True)

    if 'final_gross' not in df.columns:
        df['final_gross'] = df['to_be_paid'] + df['total_cash']

    top3 = df.nlargest(3, 'final_gross')[['driver', 'to_be_paid', 'total_cash', 'final_gross']].copy()
    top3['final_gross_display'] = top3['final_gross'].map(lambda x: f"{x:,.2f} RON")

    return render_template('summary.html',
                           week=week,
                           summary=df.to_dict(orient='records'),
                           top3=top3.to_dict(orient='records'))

@app.route('/download_all/<week>')
@login_required
def download_all(week):
    zip_path = os.path.join('uploads', 'history', week, 'drivers.zip')
    if not os.path.exists(zip_path):
        return f"No ZIP archive found for week {week}", 404
    return send_from_directory(os.path.dirname(zip_path), 'drivers.zip', as_attachment=True)

@app.route('/download_history/<week>/<filename>')
@login_required
def download_history(week, filename):
    path = os.path.join('uploads', 'history', week)
    return send_from_directory(path, filename, as_attachment=True)

@app.route('/history')
@login_required
def history():
    base = os.path.join('uploads', 'history')
    weeks = []
    if os.path.exists(base):
        for week in sorted(os.listdir(base), reverse=True):
            meta = os.path.join(base, week, 'metadata.json')
            if os.path.exists(meta):
                with open(meta) as f:
                    weeks.append(json.load(f))
    return render_template('history.html', weeks=weeks)

if __name__ == '__main__':
    os.makedirs('uploads/history', exist_ok=True)
    app.run(debug=True)